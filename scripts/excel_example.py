import openpyxl

def get_values(xl_file, name_and_cols, transformer):
    values = set()
    wb = openpyxl.load_workbook(xl_file)
    for (n, col) in name_and_cols:
        sheet = wb[n]
        rows = sheet.max_row
        for i in range(1, rows):
            raw = sheet.cell(row=i, column=col).value
            value = transformer(raw) 
            if value != None:
                values.add(value)
    return values

def validate_number(n):
    if n == None:
        return n
    n = str(n).strip()
    if len(n) == 11:
        return n
    else:
        return None

def validate_name(n):
    if n != None:
        names = n.split(None)
        if len(names) > 1:
            names.sort()
        return '-'.join(names)
    else:
        return None

user_phones = get_values('chengjiao.xlsx', [('Sheet1', 4)], validate_number)
visited_phones = get_values('visit2.xlsx', [('渠道', 3), ('中介', 4)], validate_number)

user_names = get_values('chengjiao.xlsx', [('Sheet1', 3)], validate_name)
visited_names = get_values('visit2.xlsx', [('渠道', 2), ('中介', 3)], validate_name)

"""
with open('name.txt', mode='w', encoding='utf-8') as f:
    for n in names:
        f.write(f"{n}\r\n")
"""
phones = user_phones & visited_phones
print(f"common phones {len(phones)}")

names = user_names & visited_names
print(f"common names {len(names)}")
